from __future__ import annotations

import argparse
import csv
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


PROJECT = Path.cwd()
DATA = PROJECT / "external_bulk" / "Skerget_NG2024_CoMMpass_public_supplement"
OUT = PROJECT / "analysis" / "skerget_ng2024_public_supplement"
REPORTS = PROJECT / "reports" / "validation"
SCORES = PROJECT / "analysis" / "commppass_gdc_validation" / "commppass_axis_clinical_scores.tsv"


@dataclass(frozen=True)
class Supplement:
    table: str
    title: str
    file_name: str
    url: str
    priority: str


BASE = (
    "https://static-content.springer.com/esm/art%3A10.1038%2Fs41588-024-01853-0/"
    "MediaObjects/"
)

SUPPLEMENTS = [
    Supplement(
        "Supplementary Table 1",
        "Individual patient features and data dictionary",
        "41588_2024_1853_MOESM4_ESM.xlsx",
        BASE + "41588_2024_1853_MOESM4_ESM.xlsx",
        "core",
    ),
    Supplement(
        "Supplementary Table 2",
        "Somatic SNV/INDEL events",
        "41588_2024_1853_MOESM5_ESM.xlsx",
        BASE + "41588_2024_1853_MOESM5_ESM.xlsx",
        "extended",
    ),
    Supplement(
        "Supplementary Table 3",
        "Somatic structural events",
        "41588_2024_1853_MOESM6_ESM.xlsx",
        BASE + "41588_2024_1853_MOESM6_ESM.xlsx",
        "extended",
    ),
    Supplement(
        "Supplementary Table 4",
        "Expression and fusion matrix",
        "41588_2024_1853_MOESM7_ESM.xlsx",
        BASE + "41588_2024_1853_MOESM7_ESM.xlsx",
        "large",
    ),
    Supplement(
        "Supplementary Table 5",
        "Copy-number and allele-frequency matrix",
        "41588_2024_1853_MOESM8_ESM.xlsx",
        BASE + "41588_2024_1853_MOESM8_ESM.xlsx",
        "large",
    ),
    Supplement(
        "Supplementary Table 6",
        "Gene-level LOF/GOF states",
        "41588_2024_1853_MOESM9_ESM.xlsx",
        BASE + "41588_2024_1853_MOESM9_ESM.xlsx",
        "extended",
    ),
    Supplement(
        "Supplementary Table 7",
        "Gene-expression subtype classifier results",
        "41588_2024_1853_MOESM10_ESM.xlsx",
        BASE + "41588_2024_1853_MOESM10_ESM.xlsx",
        "core",
    ),
]

FIELD_PATTERNS = [
    "patient",
    "sample",
    "iss",
    "riss",
    "albumin",
    "b2m",
    "beta",
    "ldh",
    "cytogenetic",
    "risk",
    "1q",
    "17p",
    "13q",
    "hyperdiploid",
    "subtype",
    "ccnd",
    "maf",
    "whsc1",
    "tp53",
    "kras",
    "nras",
    "braf",
    "dis3",
    "fam46c",
    "tent5c",
]


def selected_supplements(download_set: str) -> list[Supplement]:
    if download_set == "core":
        return [s for s in SUPPLEMENTS if s.priority == "core"]
    if download_set == "no-large":
        return [s for s in SUPPLEMENTS if s.priority in {"core", "extended"}]
    return SUPPLEMENTS


def run_curl(url: str, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    command = [
        "curl.exe",
        "-L",
        "--retry",
        "5",
        "--connect-timeout",
        "30",
        "-C",
        "-",
        "-o",
        str(target),
        url,
    ]
    subprocess.run(command, check=True)


def download_files(download_set: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for sup in selected_supplements(download_set):
        path = DATA / sup.file_name
        status = "present"
        if not path.exists() or path.stat().st_size == 0:
            status = "downloaded"
            run_curl(sup.url, path)
        rows.append(
            {
                "table": sup.table,
                "title": sup.title,
                "priority": sup.priority,
                "file_name": sup.file_name,
                "path": str(path),
                "size_bytes": str(path.stat().st_size if path.exists() else 0),
                "status": status,
                "url": sup.url,
            }
        )
    write_tsv(OUT / "download_manifest.tsv", rows)
    return rows


def write_tsv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def workbook_inventory(audit_large: bool = False) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for sup in SUPPLEMENTS:
        path = DATA / sup.file_name
        if not path.exists():
            continue
        size_mb = path.stat().st_size / 1024 / 1024
        if size_mb > 150 and not audit_large:
            rows.append(
                {
                    "table": sup.table,
                    "file_name": sup.file_name,
                    "size_mb": round(size_mb, 3),
                    "sheet": "",
                    "max_row": "",
                    "max_column": "",
                    "first_nonempty_row": "",
                    "status": "large_file_not_loaded",
                }
            )
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                first_nonempty = ""
                for row in ws.iter_rows(max_row=8, values_only=True):
                    values = [str(v).strip() for v in row if v not in (None, "")]
                    if values:
                        first_nonempty = " | ".join(values[:12])
                        break
                rows.append(
                    {
                        "table": sup.table,
                        "file_name": sup.file_name,
                        "size_mb": round(size_mb, 3),
                        "sheet": ws.title,
                        "max_row": ws.max_row,
                        "max_column": ws.max_column,
                        "first_nonempty_row": first_nonempty,
                        "status": "loaded",
                    }
                )
            wb.close()
        except Exception as exc:  # noqa: BLE001
            rows.append(
                {
                    "table": sup.table,
                    "file_name": sup.file_name,
                    "size_mb": round(size_mb, 3),
                    "sheet": "",
                    "max_row": "",
                    "max_column": "",
                    "first_nonempty_row": "",
                    "status": f"error: {exc}",
                }
            )
    write_tsv(OUT / "workbook_inventory.tsv", rows)
    return rows


def normalize_id(value: object) -> str:
    text = str(value).strip().upper()
    text = text.replace("-", "_")
    match = re.search(r"MMRF[_ ]?(\d+)", text)
    if match:
        return f"MMRF_{match.group(1)}"
    match = re.search(r"MMRF(\d+)", text)
    if match:
        return f"MMRF_{match.group(1)}"
    return text


def find_header_row(path: Path, sheet: str) -> int:
    preview = pd.read_excel(path, sheet_name=sheet, nrows=20, header=None)
    for idx, row in preview.iterrows():
        values = [str(v).strip().lower() for v in row.tolist() if pd.notna(v)]
        if any("patient" in v or "sample" in v for v in values):
            return int(idx)
    return 0


def read_small_sheet(path: Path, sheet: str) -> pd.DataFrame:
    header = find_header_row(path, sheet)
    return pd.read_excel(path, sheet_name=sheet, header=header)


def audit_id_matching() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if not SCORES.exists():
        write_tsv(OUT / "id_match_summary.tsv", rows)
        return rows
    scores = pd.read_csv(SCORES, sep="\t", dtype=str)
    local_ids = set()
    for col in ["case_submitter_id", "sample_submitter_id", "aliquot_submitter_id"]:
        if col in scores.columns:
            local_ids.update(scores[col].dropna().map(normalize_id))
    for sup in [s for s in SUPPLEMENTS if s.priority == "core"]:
        path = DATA / sup.file_name
        if not path.exists():
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = [ws.title for ws in wb.worksheets]
            wb.close()
        except Exception:
            continue
        for sheet in sheets:
            try:
                df = read_small_sheet(path, sheet)
            except Exception as exc:  # noqa: BLE001
                rows.append(
                    {
                        "table": sup.table,
                        "sheet": sheet,
                        "id_column": "",
                        "rows": "",
                        "unique_ids": "",
                        "local_id_overlap": "",
                        "candidate_fields": "",
                        "status": f"read_error: {exc}",
                    }
                )
                continue
            cols = [str(c) for c in df.columns]
            id_cols = [c for c in cols if re.search(r"(patient|sample|identifier|id)", c, flags=re.I)]
            candidate_fields = [
                c
                for c in cols
                if any(pattern in c.lower().replace(" ", "_") for pattern in FIELD_PATTERNS)
            ]
            if not id_cols:
                rows.append(
                    {
                        "table": sup.table,
                        "sheet": sheet,
                        "id_column": "",
                        "rows": len(df),
                        "unique_ids": "",
                        "local_id_overlap": "",
                        "candidate_fields": "; ".join(candidate_fields[:80]),
                        "status": "no_id_column_detected",
                    }
                )
                continue
            for col in id_cols:
                ids = set(df[col].dropna().map(normalize_id))
                rows.append(
                    {
                        "table": sup.table,
                        "sheet": sheet,
                        "id_column": col,
                        "rows": len(df),
                        "unique_ids": len(ids),
                        "local_id_overlap": len(ids & local_ids),
                        "candidate_fields": "; ".join(candidate_fields[:80]),
                        "status": "ok",
                    }
                )
    write_tsv(OUT / "id_match_summary.tsv", rows)
    return rows


def write_report(manifest: list[dict[str, str]], inventory: list[dict[str, object]], matches: list[dict[str, object]]) -> None:
    downloaded = [row for row in manifest if int(row["size_bytes"]) > 0]
    overlap_rows = [
        row
        for row in matches
        if str(row.get("local_id_overlap", "")).isdigit() and int(str(row["local_id_overlap"])) > 0
    ]
    lines = [
        "# Skerget NG2024 Public CoMMpass Supplement Audit",
        "",
        "Date: 2026-04-30",
        "",
        "## Download Status",
        "",
        f"- Files present/downloaded in this run: {len(downloaded)}.",
        f"- Download directory: `{DATA}`.",
        f"- Manifest: `{OUT / 'download_manifest.tsv'}`.",
        "",
        "## Workbook Inventory",
        "",
        f"- Workbook inventory: `{OUT / 'workbook_inventory.tsv'}`.",
        "- Large workbooks above 150 MB are recorded but not loaded unless `--audit-large` is used.",
        "",
        "## ID Matching",
        "",
        f"- ID match summary: `{OUT / 'id_match_summary.tsv'}`.",
    ]
    if overlap_rows:
        best = max(overlap_rows, key=lambda row: int(str(row["local_id_overlap"])))
        lines.extend(
            [
                f"- Best overlap: {best['table']} / {best['sheet']} / `{best['id_column']}` = {best['local_id_overlap']} local CoMMpass/GDC IDs.",
                "- This means patient-level molecular annotation can likely be joined to the current CoMMpass axis score table.",
            ]
        )
    else:
        lines.extend(
            [
                "- No positive overlap was detected in the audited core sheets yet.",
                "- This may mean the ID field uses a different format or that manual mapping is needed.",
            ]
        )
    lines.extend(
        [
            "",
            "## Claim Boundary",
            "",
            "- These public files can support molecular-risk and subtype annotation after verified ID matching.",
            "- They should not be used to claim PFS, treatment-response, or therapy-line validation unless those fields are explicitly found and validated.",
            "",
            "## Source",
            "",
            "- Skerget et al., Nature Genetics 2024: https://www.nature.com/articles/s41588-024-01853-0",
        ]
    )
    REPORTS.mkdir(parents=True, exist_ok=True)
    (REPORTS / "SKERGET_NG2024_PUBLIC_SUPPLEMENT_AUDIT.md").write_text(
        "\n".join(lines) + "\n", encoding="utf-8"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--download-set",
        choices=["core", "no-large", "all"],
        default="all",
        help="core downloads Tables 1 and 7; no-large downloads core plus Tables 2, 3, 6; all downloads Tables 1-7.",
    )
    parser.add_argument("--audit-large", action="store_true", help="Load large XLSX files during workbook inventory.")
    args = parser.parse_args()

    DATA.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    REPORTS.mkdir(parents=True, exist_ok=True)

    manifest = download_files(args.download_set)
    inventory = workbook_inventory(audit_large=args.audit_large)
    matches = audit_id_matching()
    write_report(manifest, inventory, matches)

    print(f"Downloaded/registered files: {len(manifest)}")
    print(f"Manifest: {OUT / 'download_manifest.tsv'}")
    print(f"Workbook inventory: {OUT / 'workbook_inventory.tsv'}")
    print(f"ID match summary: {OUT / 'id_match_summary.tsv'}")
    print(f"Report: {REPORTS / 'SKERGET_NG2024_PUBLIC_SUPPLEMENT_AUDIT.md'}")


if __name__ == "__main__":
    main()
